# 2026-08-28 운용 중 발견된 3건의 회귀 테스트 — 고치기 전엔 실패, 고친 뒤엔 통과해야 한다
# 사용법: python3 qa_known_issues.py
"""
사용자가 실사용 중 잡은 문제를 재현 가능한 형태로 굳혀 둔다. 진단만 하고 넘어가면
다음 세션이 "고쳐졌는지" 알 수 없고, 고친 뒤에도 회귀를 못 잡는다.

지금 상태(미수정)에서는 3건 모두 FAIL 이 정상이다.
"""
import datetime, json, os, re, subprocess, sys

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, '..', '입결 및 인사이트',
                    'TongTongTong_2027학년도 수시지원의 모든 것_Final오타 수정 필요.xlsx')

def dump(expr):
    out = subprocess.run(['node', '-e',
        "global.window={};require('./data.js');"
        f"process.stdout.write(JSON.stringify({expr}))"],
        capture_output=True, text=True, cwd=HERE)
    if out.returncode != 0:
        raise SystemExit('node 덤프 실패\n' + out.stderr)
    return json.loads(out.stdout)

fails = []

# ---------------------------------------------------------------- ① 복수 숫자 모집인원
# 원본 엑셀 col8 에 "인:80\n자:40" 처럼 숫자가 둘 이상 든 행이 있고, num() 이 첫 숫자만 읽어
# 나머지가 사라진다(원본 14행 중 7행은 교정 완료, 미교정 6행 58명). 엑셀이 있을 때만 검사한다.
print('=== ① 모집인원 복수 숫자 (원본 대조) ===')
if not os.path.exists(XLSX):
    print('  - 원천 엑셀 없음 → SKIP')
else:
    import openpyxl
    # ⚠️ 원본·산출물 모두 학과명에 줄바꿈이 섞여 있어 그대로 비교하면 매칭이 실패한다
    #    (실측: 한양대(ERICA) 국방지능정보융합 1행을 놓쳤다). 공백을 지우고 맞춘다.
    norm = lambda x: re.sub(r'\s+', '', str(x or ''))
    ws = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)['전체']
    raw = {}
    for r in ws.iter_rows(min_row=4, values_only=True):
        v = r[8]
        if v is None or isinstance(v, (int, float)):
            continue
        nums = re.findall(r'\d+', str(v).replace(',', ''))
        if len(nums) > 1:
            raw[(norm(r[2]), norm(r[4]), norm(r[5]), norm(r[6]))] = \
                (sum(int(x) for x in nums), int(nums[0]), str(v).replace('\n', ' / '))
    print(f'  원본에서 숫자 2개 이상인 행: {len(raw)}건')
    D = dump('window.IPSI')
    S, dc = D['schema'], D['dicts']
    iu, idp, ijt, ijn, ie = (S.index(x) for x in ('uni', 'dept', 'jhtype', 'jhname', 'enroll'))
    built = {}
    for r in D['rows']:
        built[(norm(dc['uni'][r[iu]]), norm(dc['dept'][r[idp]]), norm(r[ijt]), norm(dc['jhname'][r[ijn]]))] = r[ie]
    # ⚠️ 원본만 보고 규모를 재면 안 된다 — data_corrections.json 이 이미 절반을 잡아 뒀다.
    #    (실측: 14행 중 7행은 교정 완료. 원본만 보고 '14행 113명'이라 오판한 적이 있다.)
    bad, already, nomatch = [], [], []
    for k, (total, first, txt) in raw.items():
        got = built.get(k)
        if got is None:
            nomatch.append((k, txt, total))
        elif got == first and total != first:
            bad.append((k, txt, first, total))
        else:
            already.append((k, txt, got, total))
    print(f'  이미 교정됨 {len(already)}행 · 미교정 {len(bad)}행 · 키 매칭 실패 {len(nomatch)}행')
    for k, txt, tot in nomatch:
        print(f'  ? {k[0][:12]} {k[1][:20]} [{txt[:20]}] 합 {tot} — 키 매칭 실패, 수동 확인')
    for (u, d, t, j), txt, first, total in bad:
        print(f'  ✗ {u} {d[:18]} [{t}] 원문[{txt[:24]}] → {first} (정답 후보 {total})')
    if bad:
        fails.append(f'복수 숫자 모집인원 {len(bad)}행 — 첫 숫자만 반영됨 '
                     f'(누락 {sum(t - f for _, _, f, t in bad)}명)')
    else:
        print('  ✓ 복수 숫자 행이 전부 처리됨')

# ---------------------------------------------------------------- ② 약칭 검색
# ⚠️ 2026-08-28 교훈 — 이 검사를 처음 짤 때 app.js 의 expandToken() 을 재현하지 않고
#    원문 매칭만 해서 '진주교대·서울여대 0건'이라는 **가짜 결함**을 만들었다.
#    실제 앱은 '교대$→교육대', '여대→여자' 확장을 이미 하고 있었다.
#    그래서 규칙을 하드코딩하지 않고 **app.js 에서 함수 본문을 그대로 떠다 실행**한다.
#    앱이 규칙을 고치면 이 검사도 자동으로 따라간다.
print('\n=== ② 약칭 검색 ===')
D = dump('window.IPSI')
S, dc = D['schema'], D['dicts']
iu, idp, ijn, ir, ijt, isg, ig = (S.index(x) for x in
                                 ('uni', 'dept', 'jhname', 'region', 'jhtype', 'sigun', 'gye'))
def hay(r):
    return (dc['uni'][r[iu]] + ' ' + dc['dept'][r[idp]] + ' ' + dc['jhname'][r[ijn]] + ' ' +
            dc['region'][r[ir]] + ' ' + r[ijt] + ' ' + dc['sigun'][r[isg]] + ' ' + r[ig]).lower()
HAYS = [hay(r) for r in D['rows']]
APP_SRC = open(os.path.join(HERE, 'app.js'), encoding='utf-8').read()
_m = re.search(r'const expandToken = (t => t[\s\S]*?);\n', APP_SRC)
if not _m:
    raise SystemExit('app.js 에서 expandToken 을 찾지 못했다 — 검색 확장 규칙 위치가 바뀌었는지 확인하라')
EXPAND_SRC = _m.group(1)

def expand_many(tokens):
    """app.js 의 expandToken 을 node 로 실제 실행해 확장 결과를 받는다."""
    js = f'const f = {EXPAND_SRC}; process.stdout.write(JSON.stringify({json.dumps(tokens)}.map(f)))'
    out = subprocess.run(['node', '-e', js], capture_output=True, text=True, cwd=HERE)
    if out.returncode != 0:
        raise SystemExit('expandToken 실행 실패\n' + out.stderr)
    return json.loads(out.stdout)

def count(q, expanded):
    return sum(1 for h in HAYS if all(t in h for t in expanded))

ALIASES = ['진주교대', '서울교대', '경인교대', '서울여대', '숙명여대', '이대',
           '한국외대', '카이스트', '포스텍', '서울과기대', '한기대',
           # 분캠·과기원 한글 음차 — 원천이 영문이라 한글로는 0건이었다(2026-08-29)
           '에리카', '와이즈', '유니스트', '지스트', '디지스트', '켄텍']
miss = []
_exp = expand_many([a.lower() for a in ALIASES])
for q, e in zip(ALIASES, _exp):
    n = count(q, [e])
    print(f'  {"✓" if n else "✗"} "{q}" → {n}건  (확장: {e})')
    if not n:
        miss.append(q)
if miss:
    fails.append(f'약칭 검색 {len(miss)}건 0건 — {miss}')

# ---------------------------------------------------------------- ③ 대학 패널 캠퍼스 구분
# uniPanel 이 r.uni 로만 묶어 경북대 대구·상주가 섞인다. app.js 소스로 판정한다.
print('\n=== ③ 대학 패널 캠퍼스 구분 ===')
APP = open(os.path.join(HERE, 'app.js'), encoding='utf-8').read()
m = re.search(r'box\.innerHTML = \w+\.map\(\w+ => uniPanelHTML\(\w+, FILTERED\.filter\(([^)]*)\)', APP)
if m and 'campus' not in m.group(1):
    print(f'  ✗ 패널이 캠퍼스를 무시하고 묶는다 — filter({m.group(1).strip()})')
    knu = [r for r in D['rows'] if dc['uni'][r[iu]] == '경북대학교']
    sang = [r for r in knu if dc['sigun'][r[isg]] == '상주']
    print(f'     경북대 {len(knu)}행 중 상주 {len(sang)}행이 대구캠과 한 덩어리로 표시됨')
    fails.append('대학 패널이 캠퍼스를 구분하지 않음(경북대 대구·상주 혼재)')
elif m:
    print('  ✓ 패널이 캠퍼스를 구분한다')
else:
    print('  ? uniPanel 구성 코드를 찾지 못함 — 구조 변경 확인 필요')


# ---------------------------------------------------------------- ④ 추합 단일 오염
# 대학마다 '충원' 정의가 달라(명/배/%) 계통 차이는 손대지 않는다(checklist 🔒).
# 하지만 **같은 행 3개년 중 하나만 자릿수가 튀는 것**은 정의 차이로 설명되지 않는다 —
# 옆 셀 값이 흘러든 오염이다. 실측: 동아대 사회학 [7, 972, 4] 1행(모집 6명·지원 42명).
print('\n=== ④ 추합 단일 오염 (3개년 중 하나만 자릿수 이탈) ===')
_CH = [S.index(k) for k in ('chung26', 'chung25', 'chung24')]
_out = []
for r in D['rows']:
    vals = []
    for ci in _CH:
        t = str(r[ci] or '')
        if re.fullmatch(r'\d+', t): vals.append(int(t))
    if len(vals) < 3: continue
    mx = max(vals); others = [v for v in vals if v != mx]
    if not others: continue
    om = max(others)
    if mx >= 100 and om > 0 and mx / om >= 15:
        _out.append((mx / om, dc['uni'][r[iu]], dc['dept'][r[idp]][:18], vals))
for ratio, u, d, vals in sorted(_out, reverse=True):
    print(f'  ✗ {u} {d} 3개년 {vals} → {ratio:.0f}배 이탈')
if _out:
    fails.append(f'추합 단일 오염 {len(_out)}행 — data_corrections.json "chung" 채널로 처리하라')
else:
    print('  ✓ 3개년이 함께 크거나 함께 작다(계통 차이만 남음)')


# ---------------------------------------------------------------- ⑤ 원천 컬럼 미사용
# 엑셀 35열 중 build_data.py 가 **아예 읽지 않는 열**이 있는지 본다.
# 실측(2026-08-28): col13 필요서류·col14 복수지원·col30 2024기준 3열이 100% 채워져 있는데
# 통째로 버려지고 있었다. 특히 std24 부재로 3개년 입결 추이가 기준 다른 값을 이어 그렸다.
# 새 엑셀에서 열이 늘거나 매핑이 빠지면 여기서 잡힌다.
print('\n=== ⑤ 원천 컬럼 사용 여부 ===')
if not os.path.exists(XLSX):
    print('  - 원천 엑셀 없음 → SKIP')
else:
    import openpyxl as _ox
    _ws = _ox.load_workbook(XLSX, read_only=True, data_only=True)['전체']
    _hdr = list(next(_ws.iter_rows(min_row=3, max_row=3, values_only=True)))
    _src = open(os.path.join(HERE, 'build_data.py'), encoding='utf-8').read()
    _used = {int(x) for x in re.findall(r'r\[(\d+)\]', _src)}
    _unused = [i for i in range(len(_hdr)) if i not in _used]
    # 채워진 비율을 함께 재서 '비어 있는 열'과 '버려진 열'을 구분한다
    _rows = [r for r in _ws.iter_rows(min_row=4, values_only=True) if r[2]]
    _real = []
    for i in _unused:
        n = sum(1 for r in _rows if i < len(r) and r[i] is not None and str(r[i]).strip() not in ('', '-'))
        pct = n / len(_rows) * 100 if _rows else 0
        name = str(_hdr[i] or '').replace('\n', ' ')[:20]
        print(f'  {"✗" if pct >= 5 else "·"} col{i:2d} {name:22s} 채움 {pct:5.1f}%')
        if pct >= 5: _real.append((i, name, pct))
    if _real:
        fails.append(f'원천 컬럼 {len(_real)}열이 미사용 — {[f"col{i}({n})" for i, n, _ in _real]}')
    else:
        print(f'  ✓ 미사용 {len(_unused)}열은 전부 비어 있음(수집 대상 아님)')


# ---------------------------------------------------------------- ⑥ 3개년 세트의 연도 편식
# ⑤(미사용 열)로는 **std24 형 사고를 못 잡는다** — 읽기는 읽으니 ⑤는 통과한다.
# 진짜 원인은 따로였다: 2026-07-15 e60d82e 에서 '기준이 다르면 추세로 읽으면 안 된다'는
# 원리를 알아내고 std25 를 도입했는데, 그 커밋이 고친 건 **유불리 판정(2026 vs 2025 프레임)**
# 뿐이었다. 같은 입결을 3개년으로 그리는 화면(모달 표·스파크라인·차트B)은 점검하지 않아
# std24 가 40여 일간 비어 있었다.
# → 3개년 세트 필드인데 **특정 연도만 참조가 0**이면 그 연도가 소비처에서 누락된 것이다.
#   (배열로 묶어 쓰는 필드는 세 연도가 모두 0 — 균등하므로 정상으로 본다.)
print('\n=== ⑥ 3개년 세트의 연도 편식 ===')
_app = open(os.path.join(HERE, 'app.js'), encoding='utf-8').read()
_odd = []
for base in ('c', 'g', 'v', 'chung', 'std'):
    cnt = {y: len(re.findall(rf'\b(?:r\.)?{base}{y}\b', _app)) for y in ('26', '25', '24')}
    zero = [y for y, n in cnt.items() if n == 0]
    if zero and len(zero) < 3:
        _odd.append((base, cnt, zero))
        print(f'  ✗ {base}: ' + ' '.join(f'{base}{y}={n}' for y, n in cnt.items())
              + f' → {zero} 만 소비처 없음')
    else:
        print(f'  ✓ {base}: ' + ' '.join(f'{base}{y}={n}' for y, n in cnt.items())
              + ('  (배열 참조 — 균등)' if len(zero) == 3 else ''))
if _odd:
    fails.append(f'3개년 세트 연도 편식 {len(_odd)}건 — {[b for b, _, _ in _odd]}: '
                 '한 연도만 화면에서 안 쓰인다(std24 형 사고)')


# ---------------------------------------------------------------- ⑦ 원서접수 일정(apply_dates.js)
# ⚠️ 이 산출물은 **어떤 하네스도 보지 않는 사각지대**였다. 실측(2026-08-29): 8/11 수집본에
# 서울여대 마감이 '09-11T06:00'(새벽 6시)로 들어 있었다 — 12시간제 파싱 오류다. 접수 마감
# 시각은 학생이 원서를 넣는 마지막 순간을 좌우하므로 틀리면 곧바로 지원 실패다.
# 또한 18일 만에 7개교 일정이 바뀌었다(공주교대는 마감이 17:00→16:00 으로 앞당겨짐) —
# '접수 주간 직전 1회'로는 부족하다. 접수 D-30 부터는 주 1회 재수집하라.
print('\n=== ⑦ 원서접수 일정 불변식 ===')
_ap = os.path.join(HERE, 'apply_dates.js')
if not os.path.exists(_ap):
    print('  - apply_dates.js 없음 → SKIP')
else:
    _A = dump("(function(){require('./apply_dates.js');return window.IPSI_APPLY})()")
    _ks = [k for k in _A if k != 'meta']
    _bad = []
    for k in _ks:
        v = _A.get(k) or {}
        fr, to = v.get('from'), v.get('to')
        if not to:
            continue
        hh = int(to[11:13])
        if hh < 8:                      # 새벽 마감 = 12시간제 파싱 오류 신호
            _bad.append((k, to, f'마감이 새벽 {to[11:16]} — 12시간제 파싱 오류 의심'))
        if datetime.date.fromisoformat(to[:10]).weekday() >= 5:
            _bad.append((k, to, '마감이 주말'))
        if fr and fr > to:
            _bad.append((k, to, f'시작({fr})이 마감보다 늦다'))
    print(f'  수록 {len(_ks)}교 · 이상 {len(_bad)}건')
    for k, t, why in _bad[:12]:
        print(f'  ✗ {k} {t} — {why}')
    if _bad:
        fails.append(f'원서접수 일정 이상 {len(_bad)}건 — python3 fetch_apply_dates.py 로 재수집하라')
    # 접수 기간 범위 — 값이 통째로 옮겨져도 형식 검사만으로는 안 걸린다.
    # 2027 수시 원서접수는 9/7(월)~9/11(금). 과기원 등 자체 일정은 예외로 둔다.
    _EXEMPT = {'KAIST', 'DGIST', 'UNIST', 'GIST', 'KENTECH'}
    _oor = []
    for k in _ks:
        v = _A.get(k) or {}
        fr, to = v.get('from'), v.get('to')
        if not to or k in _EXEMPT:
            continue
        if not ('2026-09-07' <= to[:10] <= '2026-09-11'):
            _oor.append((k, to, '마감이 접수 기간(9/7~9/11) 밖'))
        if fr and not ('2026-09-01' <= fr[:10] <= '2026-09-09'):
            _oor.append((k, fr, '시작이 9/1~9/9 밖'))
    for k, t, why in _oor[:8]:
        print(f'  ✗ {k} {t} — {why}')
    if _oor:
        fails.append(f'접수 일정 범위 이탈 {len(_oor)}건 — 재수집하거나 예외로 등재하라')

    # 커버리지 방향 — data.js 에 있는 대학이 apply_dates 에서 빠지면 그 지원자는 안내를 못 본다
    _uni = dump("(function(){return window.IPSI.dicts.uni})()")
    _nomap = [u for u in _uni if u not in _A]
    if _nomap:
        print(f'  ✗ 접수일 누락 대학 {len(_nomap)}교: {_nomap[:6]}')
        fails.append(f'apply_dates 에 없는 대학 {len(_nomap)}교 — python3 fetch_apply_dates.py 재실행')
    else:
        print(f'  ✓ data.js 대학 {len(_uni)}교 전부 수록')

    # 신선도 — 접수가 다가오는데 수집본이 낡았으면 경고(실패는 아님)
    # ⚠️ 파일 mtime 을 쓰면 안 된다. 변이 테스트나 포맷 정리가 파일을 다시 쓰면 mtime 이 갱신되고,
    #    git clone 을 새로 하면 모든 파일의 mtime 이 clone 시각이 되어 '항상 0일'로 보고한다.
    #    실제로 2026-09-01 에 apply_dates.js 가 '수집 후 1일'로 나왔지만 마지막 수집은 3일 전이었다.
    #    수집 시점의 진실은 git 이 안다 — 마지막 커밋 날짜를 쓴다(git 이 없으면 mtime 으로 후퇴).
    def _collected_at(path):
        try:
            import subprocess
            out = subprocess.run(['git', 'log', '-1', '--format=%cI', '--', os.path.basename(path)],
                                 cwd=os.path.dirname(os.path.abspath(path)),
                                 capture_output=True, text=True, timeout=10).stdout.strip()
            if out:
                return datetime.date.fromisoformat(out[:10])
        except Exception:
            pass
        return datetime.date.fromtimestamp(os.path.getmtime(path))

    _age = (datetime.date.today() - _collected_at(_ap)).days
    _dday = (datetime.date(2026, 9, 7) - datetime.date.today()).days
    print(f'  수집 후 {_age}일 경과 · 접수 시작까지 {_dday}일')
    if _dday <= 30 and _age >= 7:
        print(f'  ⚠ 접수 D-{_dday} 인데 수집본이 {_age}일 됐다 — python3 fetch_apply_dates.py 재실행 권장')


# ---------------------------------------------------------------- ⑧ 대학별 행수 래칫
# ⚠️ 실측(2026-08-29): 경북대 680행을 통째로 지워도 verify_data·qa_comp_ratio·qa_chungwon·
# qa_known_issues 가 **전부 통과**했다. meta.nRows 만 함께 줄이면 아무 불변식도 걸리지 않는다.
# '없는 것'은 화면에 흔적을 남기지 않으므로 학생은 그 학과가 2027에 폐지된 줄 안다.
# 엑셀 갱신 때 시트 파싱이 끊기거나 한 대학 블록이 누락되는 사고를 여기서 잡는다.
# 기준선은 row_baseline.json — 의도한 증감이면 --save-baseline 으로 갱신한다.
print('\n=== ⑧ 대학별 행수 래칫 ===')
_bl_path = os.path.join(HERE, 'row_baseline.json')
_cur = dump("(function(){const D=window.IPSI,i=D.schema.indexOf('uni'),c={};"
            "for(const r of D.rows){const u=D.dicts.uni[r[i]];c[u]=(c[u]||0)+1}"
            "return {nRows:D.rows.length,nUni:Object.keys(c).length,perUni:c}})()")
if '--save-baseline' in sys.argv:
    json.dump(_cur, open(_bl_path, 'w', encoding='utf-8'), ensure_ascii=False)
    print(f"  기준선 갱신: {_cur['nRows']}행 · {_cur['nUni']}교")
elif not os.path.exists(_bl_path):
    print('  - row_baseline.json 없음 → --save-baseline 으로 생성하라')
else:
    _bl = json.load(open(_bl_path, encoding='utf-8'))
    _gone = sorted(set(_bl['perUni']) - set(_cur['perUni']))
    _new = sorted(set(_cur['perUni']) - set(_bl['perUni']))
    # 대학이 통째로 사라지는 것은 언제나 사고다. 행수는 10% 이상 급변만 본다.
    _shrunk = [(u, _bl['perUni'][u], _cur['perUni'][u]) for u in _bl['perUni']
               if u in _cur['perUni']
               and _cur['perUni'][u] < _bl['perUni'][u] * 0.9]
    print(f"  {_cur['nRows']}행 · {_cur['nUni']}교  (기준선 {_bl['nRows']}행 · {_bl['nUni']}교)")
    for u in _gone:
        print(f"  ✗ 대학 소멸: {u} (기준선 {_bl['perUni'][u]}행)")
    for u, a, b in _shrunk:
        print(f"  ✗ 행수 급감: {u} {a} → {b} ({(b/a-1)*100:.0f}%)")
    for u in _new:
        print(f"  + 신규 대학: {u} ({_cur['perUni'][u]}행)")
    if _gone or _shrunk:
        fails.append(f"대학 소멸 {len(_gone)}교 · 행수 급감 {len(_shrunk)}교 — "
                     f"엑셀 파싱이 끊겼는지 확인하라. 의도한 변경이면 "
                     f"python3 qa_known_issues.py --save-baseline")
    elif not _new:
        print('  ✓ 소멸·급감 없음')


# ---------------------------------------------------------------- ⑨ 핵심 분포 래칫
# ⚠️ 실측(2026-08-29): 수능최저를 300행에서 지워도, 고사일을 요일만 맞춰 옮겨도,
# 카테고리를 재라벨해도 하네스 7종이 **전부 통과**했다. 값의 '형식'만 보고 '분포'는 아무도
# 안 봤기 때문이다. 삭제·대량 재라벨 계열 사고는 분포가 먼저 무너진다.
# 기준선 dist_baseline.json 대비 5% 넘게 움직이면 실패 — 의도한 변경은 --save-baseline.
print('\n=== ⑨ 핵심 분포 래칫 ===')
_db = os.path.join(HERE, 'dist_baseline.json')
_cur = dump("""(function(){const D=window.IPSI,S=D.schema,dc=D.dicts,g=n=>S.indexOf(n);
const SU=new Date(2026,10,19);
const when=t=>{const RE=/(\\d{1,2})\\.\\s*(\\d{1,2})/g;const h=[...String(t||'').matchAll(RE)];
  if(!h.length)return null;const ds=h.map(m=>new Date(2026,+m[1]-1,+m[2]));
  return ds.every(d=>d<SU)?'pre':ds.every(d=>d>SU)?'post':null};
const b={hasChoejeo1:D.rows.filter(r=>r[g('hasChoejeo')]===1).length,
 choejeoUniq:new Set(D.rows.map(r=>dc.choejeo[r[g('choejeo')]]).filter(x=>x&&x!=='없음')).size,
 dateFilled:D.rows.filter(r=>(dc.date[r[g('date')]]||'').trim()).length,
 dateUniq:new Set(D.rows.map(r=>dc.date[r[g('date')]]).filter(x=>x&&x.trim())).size,
 examPre:0,examPost:0,dkind:{},catRows:{}};
D.rows.forEach(r=>{const w=when(dc.date[r[g('date')]]);if(w==='pre')b.examPre++;else if(w==='post')b.examPost++;
 const k=r[g('dkind')]||'(빈)';b.dkind[k]=(b.dkind[k]||0)+1;
 for(const c of (r[g('cats')]||[]))b.catRows[c]=(b.catRows[c]||0)+1});
return b})()""")
if '--save-baseline' in sys.argv:
    json.dump(_cur, open(_db, 'w', encoding='utf-8'), ensure_ascii=False)
    print('  기준선 갱신')
elif not os.path.exists(_db):
    print('  - dist_baseline.json 없음 → --save-baseline 으로 생성하라')
else:
    _bl = json.load(open(_db, encoding='utf-8'))
    _off = []
    for k in ('hasChoejeo1', 'choejeoUniq', 'dateFilled', 'dateUniq', 'examPre', 'examPost'):
        a, b2 = _bl.get(k, 0), _cur.get(k, 0)
        if a and abs(b2 - a) / a > 0.05:
            _off.append((k, a, b2))
    for grp in ('dkind', 'catRows'):
        for k, a in _bl.get(grp, {}).items():
            b2 = _cur.get(grp, {}).get(k, 0)
            if a >= 20 and abs(b2 - a) / a > 0.05:
                _off.append((f'{grp}.{k}', a, b2))
        for k in set(_cur.get(grp, {})) - set(_bl.get(grp, {})):
            _off.append((f'{grp}.{k}', 0, _cur[grp][k]))
    print(f"  최저있음 {_cur['hasChoejeo1']} · 고사일 {_cur['dateFilled']} · 수능전 {_cur['examPre']} · 수능후 {_cur['examPost']}")
    for k, a, b2 in _off[:12]:
        print(f'  ✗ {k}: {a} → {b2} ({(b2/a-1)*100:+.0f}%)' if a else f'  ✗ {k}: 신규 {b2}')
    if _off:
        fails.append(f'분포 래칫 이탈 {len(_off)}건 — 삭제·대량 재라벨 사고인지 확인하라. '
                     f'의도한 변경이면 python3 qa_known_issues.py --save-baseline')
    else:
        print('  ✓ 기준선 대비 5% 이내')

# ---------------------------------------------------------------- ⑩ 수능최저 파생 정합
# hasChoejeo 는 choejeo 원문에서 파생되는 값인데 서로 대조된 적이 없었다.
print('\n=== ⑩ 수능최저 파생 정합 ===')
_mm = dump("""(function(){const D=window.IPSI,S=D.schema,dc=D.dicts,g=n=>S.indexOf(n);
const bad=[];D.rows.forEach((r,i)=>{const t=(dc.choejeo[r[g('choejeo')]]||'').trim();
 const want=(t&&t!=='없음')?1:0; if(r[g('hasChoejeo')]!==want)
   bad.push([dc.uni[r[g('uni')]],String(dc.dept[r[g('dept')]]).slice(0,12),t.slice(0,18),r[g('hasChoejeo')],want])});
return bad})()""")
print(f'  불일치 {len(_mm)}건')
for x in _mm[:6]:
    print(f'  ✗ {x[0]} {x[1]} | 최저 {x[2]!r} | hasChoejeo={x[3]} (기대 {x[4]})')
if _mm:
    fails.append(f'hasChoejeo 파생 불일치 {len(_mm)}건 — choejeo 원문과 어긋난다')


# ---------------------------------------------------------------- ⑪ index.html 자산 배선·캐시버스터
# ⚠️ 실측(2026-08-29): insights.js 스크립트 태그를 지워도 7개 하네스가 전부 통과했다.
# app.js 는 window.IPSI_INSIGHTS 가 없으면 조용히 넘어가므로 에러조차 안 난다.
# 캐시버스터가 stale 이면 접수 주간에 급히 고친 마감 시각을 배포해도 재방문자는 옛 파일을 쓴다
# — 실제로 이날 apply_dates.js 재수집본이 스탬프 없이 커밋될 뻔했다.
print('\n=== ⑪ index.html 자산 배선·캐시버스터 ===')
import hashlib
_idx_p = os.path.join(HERE, 'index.html')
_stamp = open(os.path.join(HERE, 'stamp_assets.py'), encoding='utf-8').read()
_m = re.search(r'ASSETS\s*=\s*\[([^\]]*)\]', _stamp)
_assets = re.findall(r"'([^']+)'", _m.group(1)) if _m else []
_idx_raw = open(_idx_p, encoding='utf-8').read()
# ⚠️ 주석 안의 문자열까지 세면 태그를 주석 처리한 사고를 놓친다(첫 구현이 실제로 그랬다).
_idx = re.sub(r'<!--.*?-->', '', _idx_raw, flags=re.S)
_bad = []
for a in _assets:
    ap = os.path.join(HERE, a)
    if not os.path.exists(ap):
        _bad.append(f'{a}: 파일 없음'); continue
    h = hashlib.sha1(open(ap, 'rb').read()).hexdigest()[:8]
    # 실제 <script src>/<link href> 로 실려 있는지 — 파일명만 어디에 있는 것으로는 부족하다
    tag = re.search(r'<(?:script|link)\b[^>]*\b(?:src|href)="' + re.escape(a) + r'(?:\?v=([a-f0-9]+))?"', _idx)
    mm = tag if tag and tag.group(1) else None
    if not tag:
        _bad.append(f'{a}: index.html 에 script/link 태그로 실려 있지 않다 (배선 끊김)')
    elif not mm:
        _bad.append(f'{a}: 캐시버스터(?v=)가 없다 → python3 stamp_assets.py')
    elif mm.group(1) != h:
        _bad.append(f'{a}: 캐시버스터 stale (파일 {h} / index {mm.group(1)}) → python3 stamp_assets.py')
print(f'  자산 {len(_assets)}종 검사')
for b in _bad:
    print('  ✗', b)
if _bad:
    fails.append(f'index.html 자산 배선·스탬프 이상 {len(_bad)}건')
else:
    print('  ✓ 전부 참조되고 스탬프도 최신')

# ---------------------------------------------------------------- ⑫ SCHEMA 밖 사이드맵 배선
# raw·chungDoubt 는 rows 배열이 아니라 **행 인덱스를 키로 하는 별도 맵**이라
# probe_fields.js 의 마커 주입 방식으로는 검사되지 않는다(DATA_FIELDS 에 넣을 수가 없다).
# 그래서 여기서 따로 본다 — (a) app.js 가 실제로 읽는가 (b) 인덱스가 rows 범위 안인가.
# ⚠️ 사이드맵은 행 인덱스로만 연결돼 있어, rows 를 나중에 필터·정렬하는 코드가 생기면
#    원문이 조용히 다른 행에 붙는다. 그 사고는 화면에서만 드러난다.
print()
print('=== ⑫ SCHEMA 밖 사이드맵(raw·chungDoubt) 배선 ===')
_app_src = open(os.path.join(HERE, 'app.js'), encoding='utf-8').read()
_side = {'raw': 'D.raw', 'chungDoubt': 'D.chungDoubt'}
_sbad = []
_nrows = len(D["rows"])
for _k, _ref in _side.items():
    _m = D.get(_k)
    if not _m:
        print(f'  · {_k}: data.js 에 없음 — 건너뜀')
        continue
    if _ref not in _app_src:
        _sbad.append(f'{_k}: data.js 에 {len(_m)}항목 있는데 app.js 가 {_ref} 를 읽지 않는다(배선 끊김)')
        continue
    _oob = [i for i in _m if not (0 <= int(i) < _nrows)]
    if _oob:
        _sbad.append(f'{_k}: 행 범위 밖 인덱스 {len(_oob)}개 (rows={_nrows})')
    else:
        print(f'  ✓ {_k}: {len(_m)}항목, app.js 가 읽음, 인덱스 전부 범위 내')
for _b in _sbad:
    print('  ✗', _b)
if _sbad:
    fails.append(f'사이드맵 배선 이상 {len(_sbad)}건')


# ---------------------------------------------------------------- 결론
print()
if fails:
    print(f'미해결 {len(fails)}건:')
    for f in fails:
        print('  ✗', f)
    print('\n(2026-08-28 진단 시점에는 3건 모두 미수정이 정상이다. 고친 뒤 이 스크립트가 통과해야 한다.)')
    sys.exit(1)
print('OK  알려진 3건이 모두 해소됨')
